import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Create Workbook
wb = openpyxl.Workbook()

# Setup Colors & Styles
NAVY_HEADER_FILL = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
EMERALD_ACCENT_FILL = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
CYAN_ACCENT_FILL = PatternFill(start_color="0284C7", end_color="0284C7", fill_type="solid")
LIGHT_BG_FILL = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
ZEBRA_FILL = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

FONT_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_TITLE = Font(name="Calibri", size=16, bold=True, color="0F172A")
FONT_SUBTITLE = Font(name="Calibri", size=11, italic=True, color="475569")
FONT_BOLD = Font(name="Calibri", size=10, bold=True, color="0F172A")
FONT_REGULAR = Font(name="Calibri", size=10, color="1E293B")
FONT_CODE = Font(name="Consolas", size=9, color="047857", bold=True)

THIN_BORDER = Border(
    left=Side(style='thin', color='CBD5E1'),
    right=Side(style='thin', color='CBD5E1'),
    top=Side(style='thin', color='CBD5E1'),
    bottom=Side(style='thin', color='CBD5E1')
)

ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

def style_table(ws, start_row, header_cols):
    ws.row_dimensions[start_row].height = 26
    for col_num in range(1, len(header_cols) + 1):
        cell = ws.cell(row=start_row, column=col_num)
        cell.fill = NAVY_HEADER_FILL
        cell.font = FONT_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

# ----------------------------------------------------
# SHEET 1: Executive Overview
# ----------------------------------------------------
ws1 = wb.active
ws1.title = "Executive Overview"

ws1.cell(row=1, column=1, value="PlantTwin AI — Technology Stack & Dependencies").font = FONT_TITLE
ws1.cell(row=2, column=1, value="Master Architecture, Core Technologies, and Dependencies Breakdown (v1.1 Enterprise)").font = FONT_SUBTITLE

overview_headers = ["Category", "Primary Technology", "Version / Standard", "Key Capabilities & Architectural Role"]
style_table(ws1, 4, overview_headers)
for i, h in enumerate(overview_headers, start=1):
    ws1.cell(row=4, column=i, value=h)

overview_data = [
    ("Frontend Framework", "React 18 + TypeScript", "React 18.2.0 / TS 5.2", "Single Page Application, strict type safety, component-driven UI"),
    ("Build System & Styling", "Vite + Tailwind CSS", "Vite 5.1 / Tailwind 3.4", "Sub-second HMR development, utility-first dark/light corporate styling"),
    ("Backend API Framework", "FastAPI + Python 3.10", "FastAPI 0.110.0", "Asynchronous ASGI Web API, auto OpenAPI docs, high-throughput tag ingestion"),
    ("Database Engine", "PostgreSQL + TimescaleDB", "PostgreSQL 15", "Relational metadata + hypertable time-series telemetry storage"),
    ("Caching & Message Broker", "Redis + MQTT", "Redis 5.0 / Mosquitto", "In-memory telemetry cache, real-time pub/sub messaging broker"),
    ("AI / ML Pipeline Engine", "Scikit-Learn + XGBoost + SHAP", "Scikit-Learn 1.4.1", "Equipment RUL prediction, anomaly detection, explainable AI feature attribution"),
    ("GenAI Copilot Interface", "LangChain + Google Gemini", "LangChain 0.1.13", "Natural language P&ID & telemetry query assistant with RAG context"),
    ("Industrial Protocols", "Siemens S7, OPC-UA, MQTT, Modbus", "Snap7 / OPC-UA 0.98", "Native OT PLC drivers for Siemens, Allen-Bradley, and Schneider gateways"),
    ("DevOps & Orchestration", "Docker & Docker Compose", "Compose v2", "Containerized microservice stack, single-command production deployment")
]

for row_idx, data in enumerate(overview_data, start=5):
    ws1.row_dimensions[row_idx].height = 20
    fill = ZEBRA_FILL if row_idx % 2 == 0 else LIGHT_BG_FILL
    for col_idx, val in enumerate(data, start=1):
        cell = ws1.cell(row=row_idx, column=col_idx, value=val)
        cell.fill = fill
        cell.border = THIN_BORDER
        cell.alignment = ALIGN_LEFT
        if col_idx == 1:
            cell.font = FONT_BOLD
        elif col_idx == 3:
            cell.font = FONT_CODE
            cell.alignment = ALIGN_CENTER
        else:
            cell.font = FONT_REGULAR

# ----------------------------------------------------
# SHEET 2: Backend Dependencies
# ----------------------------------------------------
ws2 = wb.create_sheet(title="Backend Dependencies")
ws2.cell(row=1, column=1, value="PlantTwin AI — Python Backend Dependencies (requirements.txt)").font = FONT_TITLE
ws2.cell(row=2, column=1, value="FastAPI, Async SQLAlchemy, AI/ML, OT Protocols, and GenAI Copilot Packages").font = FONT_SUBTITLE

backend_headers = ["Category", "Package Name", "Version Spec", "Type", "Purpose & Role in PlantTwin AI"]
style_table(ws2, 4, backend_headers)
for i, h in enumerate(backend_headers, start=1):
    ws2.cell(row=4, column=i, value=h)

backend_data = [
    # Core API & Server
    ("Core Framework", "fastapi", ">= 0.110.0", "Production", "Asynchronous Web API framework, routing, request validation"),
    ("Core Server", "uvicorn[standard]", ">= 0.28.0", "Production", "High-performance ASGI HTTP & WebSocket server"),
    ("Data Validation", "pydantic[email]", ">= 2.6.4", "Production", "Data parsing, schema enforcement, and type validation"),
    ("Settings Manager", "pydantic-settings", ">= 2.2.1", "Production", "Environment variable loading and configuration management"),
    ("Email Validation", "email-validator", ">= 2.1.0", "Production", "Email string syntax and deliverability checking"),

    # Database & ORM
    ("Database ORM", "sqlalchemy", ">= 2.0.28", "Production", "Async SQL Object Relational Mapper for PostgreSQL / SQLite"),
    ("Postgres Driver", "asyncpg", ">= 0.29.0", "Production", "Fast asynchronous PostgreSQL database driver for TimescaleDB"),
    ("SQLite Driver", "aiosqlite", ">= 0.20.0", "Production", "Asynchronous SQLite driver for local dev & testing"),
    ("DB Migration", "alembic", ">= 1.13.1", "Production", "Database schema migration tracking and revision scripts"),

    # Cache & Messaging
    ("In-Memory Cache", "redis", ">= 5.0.3", "Production", "Redis client for telemetry caching and pub/sub message distribution"),
    ("Async Task Queue", "celery", ">= 5.3.6", "Production", "Distributed task queue for background AI training & alert dispatch"),

    # Security & Auth
    ("JWT Authentication", "python-jose[cryptography]", ">= 3.3.0", "Production", "JSON Web Token encoding, decoding, and signature verification"),
    ("Password Hashing", "passlib[bcrypt]", ">= 1.7.4", "Production", "Password hashing utilities with bcrypt context"),
    ("Bcrypt Encryption", "bcrypt", ">= 4.1.2", "Production", "Cryptographic password hashing library"),
    ("Multipart Form", "python-multipart", ">= 0.0.9", "Production", "Parsing multipart/form-data for file uploads"),

    # Networking & WebSockets
    ("HTTP Client", "httpx", ">= 0.27.0", "Production", "Async HTTP client for external service integration"),
    ("WebSocket Engine", "websockets", ">= 12.0", "Production", "Real-time bi-directional telemetry streaming connection"),

    # OT Industrial Protocols
    ("Industrial OPC-UA", "opcua", ">= 0.98.13", "Production", "Native OPC-UA client/server driver for SCADA connectivity"),
    ("Industrial MQTT", "gmqtt", ">= 0.6.12", "Production", "Async MQTT client supporting Sparkplug B telemetry payloads"),
    ("Industrial Modbus", "pymodbus", ">= 3.6.8", "Production", "Modbus TCP/RTU protocol library for legacy PLC registers"),

    # Machine Learning & AI
    ("Machine Learning", "scikit-learn", ">= 1.4.1.post1", "Production", "Predictive modeling, Isolation Forest anomaly detection"),
    ("Numerical Computing", "numpy", ">= 1.26.4", "Production", "High-performance vector and array computations"),
    ("Data Analysis", "pandas", ">= 2.2.1", "Production", "Dataframe manipulation for time-series sensor logs"),

    # GenAI Copilot
    ("LLM Orchestration", "langchain", ">= 0.1.13", "Production", "GenAI agent chains, prompt templates, and memory handling"),
    ("Google Gemini GenAI", "langchain-google-genai", ">= 1.0.1", "Production", "LangChain integration for Google Gemini LLM model"),
    ("Gemini SDK", "google-generativeai", ">= 0.4.1", "Production", "Official Google Gemini SDK for industrial assistant chat"),

    # Monitoring & Logging
    ("Prometheus Metrics", "prometheus-client", ">= 0.20.0", "Production", "Exporter for application metrics and operational health"),
    ("Structured Logging", "structlog", ">= 24.1.0", "Production", "Structured JSON logging for enterprise audit trails"),

    # Dev & Testing
    ("Test Framework", "pytest", ">= 8.1.1", "Development", "Unit and integration test runner"),
    ("Async Test Plugin", "pytest-asyncio", ">= 0.23.5", "Development", "Asyncio support for pytest async test functions"),
    ("Coverage Tool", "pytest-cov", ">= 4.1.0", "Development", "Code coverage reporting plugin for pytest"),
    ("Code Formatter", "black", ">= 24.3.0", "Development", "Uncompromising Python code formatter"),
    ("Import Sorter", "isort", ">= 5.13.2", "Development", "Python import statement sorter"),
    ("Linter", "flake8", ">= 7.0.0", "Development", "PEP8 code style and syntax error checker"),
    ("Static Type Checker", "mypy", ">= 1.9.0", "Development", "Static type checker for Python codebase")
]

for row_idx, data in enumerate(backend_data, start=5):
    ws2.row_dimensions[row_idx].height = 19
    fill = ZEBRA_FILL if row_idx % 2 == 0 else LIGHT_BG_FILL
    for col_idx, val in enumerate(data, start=1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=val)
        cell.fill = fill
        cell.border = THIN_BORDER
        cell.alignment = ALIGN_LEFT
        if col_idx == 1:
            cell.font = FONT_BOLD
        elif col_idx == 2:
            cell.font = FONT_CODE
        elif col_idx == 3:
            cell.font = FONT_CODE
            cell.alignment = ALIGN_CENTER
        elif col_idx == 4:
            cell.alignment = ALIGN_CENTER
            cell.font = FONT_BOLD if val == "Production" else FONT_REGULAR
        else:
            cell.font = FONT_REGULAR

# ----------------------------------------------------
# SHEET 3: Frontend Dependencies
# ----------------------------------------------------
ws3 = wb.create_sheet(title="Frontend Dependencies")
ws3.cell(row=1, column=1, value="PlantTwin AI — React Frontend Dependencies (package.json)").font = FONT_TITLE
ws3.cell(row=2, column=1, value="React 18, TypeScript, Tailwind CSS, Recharts, and Lucide Icons").font = FONT_SUBTITLE

frontend_headers = ["Category", "Package Name", "Version Spec", "Type", "Purpose & Role in PlantTwin UI"]
style_table(ws3, 4, frontend_headers)
for i, h in enumerate(frontend_headers, start=1):
    ws3.cell(row=4, column=i, value=h)

frontend_data = [
    # Production Dependencies
    ("UI Library", "react", "^18.2.0", "Production", "Core component-driven user interface framework"),
    ("DOM Renderer", "react-dom", "^18.2.0", "Production", "React rendering package for browser DOM"),
    ("Client Routing", "react-router-dom", "^6.22.3", "Production", "Declarative routing for single-page application views"),
    ("HTTP Client", "axios", "^1.6.8", "Production", "Promise-based HTTP client for FastAPI backend calls"),
    ("Chart & Graphs", "recharts", "^2.12.3", "Production", "Composable chart library for telemetry time-series graphs"),
    ("Iconography Set", "lucide-react", "^0.359.0", "Production", "Clean, consistent SVG icon set for UI dashboards"),
    ("CSS Utility Merger", "tailwind-merge", "^2.2.2", "Production", "Utility to merge Tailwind CSS classes without conflicts"),
    ("Classnames Helper", "clsx", "^2.1.0", "Production", "Utility for constructing className strings conditionally"),

    # Dev Dependencies
    ("Build Tool & Server", "vite", "^5.1.6", "Development", "Next-generation frontend tooling & ultra-fast dev server"),
    ("TypeScript Compiler", "typescript", "^5.2.2", "Development", "Typed JavaScript language extension"),
    ("React Plugin for Vite", "@vitejs/plugin-react", "^4.2.1", "Development", "Vite plugin enabling Fast Refresh for React components"),
    ("CSS Framework", "tailwindcss", "^3.4.1", "Development", "Utility-first CSS framework for custom corporate styling"),
    ("PostCSS Processor", "postcss", "^8.4.38", "Development", "Tool for transforming CSS with JavaScript plugins"),
    ("CSS Vendor Prefixer", "autoprefixer", "^10.4.19", "Development", "PostCSS plugin to parse CSS and add vendor prefixes"),
    ("TypeScript Node Types", "@types/node", "^20.11.30", "Development", "TypeScript type definitions for Node.js APIs"),
    ("TypeScript React Types", "@types/react", "^18.2.66", "Development", "TypeScript type definitions for React core"),
    ("TypeScript DOM Types", "@types/react-dom", "^18.2.22", "Development", "TypeScript type definitions for React DOM")
]

for row_idx, data in enumerate(frontend_data, start=5):
    ws3.row_dimensions[row_idx].height = 19
    fill = ZEBRA_FILL if row_idx % 2 == 0 else LIGHT_BG_FILL
    for col_idx, val in enumerate(data, start=1):
        cell = ws3.cell(row=row_idx, column=col_idx, value=val)
        cell.fill = fill
        cell.border = THIN_BORDER
        cell.alignment = ALIGN_LEFT
        if col_idx == 1:
            cell.font = FONT_BOLD
        elif col_idx == 2:
            cell.font = FONT_CODE
        elif col_idx == 3:
            cell.font = FONT_CODE
            cell.alignment = ALIGN_CENTER
        elif col_idx == 4:
            cell.alignment = ALIGN_CENTER
            cell.font = FONT_BOLD if val == "Production" else FONT_REGULAR
        else:
            cell.font = FONT_REGULAR

# ----------------------------------------------------
# SHEET 4: Infrastructure & Hardware
# ----------------------------------------------------
ws4 = wb.create_sheet(title="Infrastructure & Hardware")
ws4.cell(row=1, column=1, value="PlantTwin AI — Infrastructure, Databases & Test Bench Hardware").font = FONT_TITLE
ws4.cell(row=2, column=1, value="Containerization, Database Extensions, Broker Infrastructure, and Physical PLC Kit").font = FONT_SUBTITLE

infra_headers = ["Layer", "Component / Hardware", "Specification", "Deployment Mode", "Operational Role"]
style_table(ws4, 4, infra_headers)
for i, h in enumerate(infra_headers, start=1):
    ws4.cell(row=4, column=i, value=h)

infra_data = [
    ("Container Engine", "Docker Engine", "v24.0+", "Cloud / Edge", "Containerized deployment environment for microservices"),
    ("Orchestration", "Docker Compose", "v2.20+", "On-Prem / Edge", "Multi-container app definition (Frontend, Backend, Postgres, Redis)"),
    ("Database Engine", "PostgreSQL", "v15.6", "Cloud / Edge", "Primary relational database for assets, users, alarms, work orders"),
    ("Time-Series Extension", "TimescaleDB", "v2.13+", "Postgres Plugin", "Hypertables optimized for high-frequency sensor telemetry ingestion"),
    ("In-Memory Cache", "Redis Server", "v7.2+", "Cloud / Edge", "Pub/sub event broker & sub-second live telemetry cache"),
    ("Industrial PLC Rig", "Siemens S7-1200 PLC", "CPU 1214C DC/DC/DC", "Physical Hardware", "Hardware test bench PLC executing S7 protocol communication"),
    ("Temperature Sensor", "Analog Temperature Rig", "0-100 °C / 4-20mA", "Physical Hardware", "Live thermal telemetry source connected to S7 PLC analog input"),
    ("Vibration Sensor", "Piezo Accelerometer", "0-50 mm/s RMS", "Physical Hardware", "Vibration data source for bearing degradation RUL AI modeling"),
    ("Power Supply", "Siemens SITOP PSU100M", "24V DC / 5A", "Physical Hardware", "Industrial power supply unit powering PLC and sensor rig"),
    ("Managed Switch", "Industrial Ethernet Switch", "5-Port 10/100 Mbps", "Physical Hardware", "Isolated OT network switch connecting PLC to edge gateway")
]

for row_idx, data in enumerate(infra_data, start=5):
    ws4.row_dimensions[row_idx].height = 19
    fill = ZEBRA_FILL if row_idx % 2 == 0 else LIGHT_BG_FILL
    for col_idx, val in enumerate(data, start=1):
        cell = ws4.cell(row=row_idx, column=col_idx, value=val)
        cell.fill = fill
        cell.border = THIN_BORDER
        cell.alignment = ALIGN_LEFT
        if col_idx == 1:
            cell.font = FONT_BOLD
        elif col_idx == 2:
            cell.font = FONT_BOLD
        elif col_idx == 3:
            cell.font = FONT_CODE
            cell.alignment = ALIGN_CENTER
        else:
            cell.font = FONT_REGULAR

# Adjust Column Widths across all sheets
for ws in [ws1, ws2, ws3, ws4]:
    ws.views.sheetView[0].showGridLines = True
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len and cell.row > 2: # Skip main title length
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

# Save Workbook
output_path = r"C:\Users\ayush\.gemini\antigravity\scratch\PlantTwin_AI_Tech_Stack_and_Dependencies.xlsx"
wb.save(output_path)
print(f"Successfully generated Excel workbook at: {output_path}")
